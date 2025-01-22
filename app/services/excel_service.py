import base64
import openpyxl
import os
import logging
from io import BytesIO
import requests
from app.services.prisma_service import fetch_template_from_prisma, get_template_from_prisma
from app.services.ypareo_service import YpareoService
from app.utils.utils import convert_minutes_to_hours_and_minutes
from prisma import Prisma
from docx import Document
from app.core.template_mapping import get_template_id_from_group_name


def download_excel_from_url(url: str) -> BytesIO:
    """
    Télécharge le fichier Excel depuis une URL et le retourne en BytesIO.
    """
    try:
        logging.info(f"Téléchargement du fichier depuis l'URL : {url}")
        response = requests.get(url)
        response.raise_for_status()
        return BytesIO(response.content)
    except Exception as e:
        logging.error(f"Erreur lors du téléchargement du fichier : {str(e)}")
        raise ValueError(f"Erreur lors du téléchargement du fichier : {str(e)}")

def copy_multiple_cells(source_url: str, template_path: str, output_dir: str) -> str:
    """
    Copie les valeurs des cellules spécifiées dans le fichier source vers les cellules correspondantes dans le template.
    """
    try:
        # Télécharger le fichier source depuis l'URL
        logging.info(f"Téléchargement du fichier source depuis l'URL : {source_url}")
        response = requests.get(source_url)
        response.raise_for_status()
        source_file = BytesIO(response.content)

        source_wb = openpyxl.load_workbook(source_file)
        template_wb = openpyxl.load_workbook(template_path)
        source_ws = source_wb.active
        template_ws = template_wb.active

        # Déterminer quel template est utilisé
        template_name = os.path.basename(template_path)
        
        # Configuration des colonnes pour chaque template
        template_configs = {
            "BG-TP-S1.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L', 'O', 'R', 'U', 'X', 'AD', 'AG', 'AJ', 'AM', 'AP', 'AV', 'AY', 'BE', 'BH', 'BK', 'BN', 'BQ', 'BT', 'BW', 'BZ'],
                "target_columns": ['B', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O', 'P', 'R', 'S', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']
            },
            "BG-TP-S2.xlsx": {
                "source_columns": ['B', 'F', 'I'],
                "target_columns": ['B', 'D', 'E']
            },
            "BG-TP-S3.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L', 'O', 'R', 'U', 'X', 'AA', 'AD', 'AG', 'AJ', 'AM', 'AP', 'AS', 'AV', 'AY', 'BB', 'BE'],
                "target_columns": ['B', 'D', 'E', 'F', 'G', 'I', 'J', 'L', 'M', 'O', 'P', 'R', 'S', 'T', 'U']
            },
            "BG-TP-S4.xlsx": {
                "source_columns": ['B', 'F', 'I'],
                "target_columns": ['B', 'D', 'E']
            },
            "BG-TP-S5.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L', 'O', 'R', 'U', 'X', 'AA', 'AD', 'AG', 'AJ', 'AM', 'AP', 'AS', 'AV', 'AY', 'BB', 'BE', 'BH', 'BK', 'BN', 'BQ'],
                "target_columns": ['B', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'N', 'P', 'Q', 'R', 'T', 'U', 'V', 'W', 'X', 'Y']
            },
            "BG-TP-S6.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L'],
                "target_columns": ['B', 'D', 'E', 'F']
            },
            
            "BG-ALT-S1.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L', 'R', 'U', 'AA', 'AG', 'AJ', 'AM', 'AP', 'AS', 'AV', 'AY', 'BB'],
                "target_columns": ['B', 'D', 'E', 'F', 'H', 'I', 'K', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
            },
            "BG-ALT-S2.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L', 'O', 'U', 'X', 'AA', 'AG', 'AM', 'AP', 'AS', 'AV', 'AY', 'BB', 'BE'],
                "target_columns": ['B', 'D', 'E', 'F', 'G', 'I', 'J', 'K', 'M', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
            },
            "BG-ALT-S3.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L', 'O', 'R', 'X', 'AA', 'AG', 'AM', 'AP', 'AS', 'AV', 'AY'],
                "target_columns": ['B', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'M', 'O', 'P', 'Q', 'R', 'S']
            },
            "BG-ALT-S4.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L', 'R', 'U', 'X', 'AA', 'AG', 'AM', 'AP', 'AS', 'AV', 'AY'],
                "target_columns": ['B', 'D', 'E', 'F', 'H', 'J', 'K', 'M', 'O', 'P', 'Q', 'R', 'S']
            },
            "BG-ALT-S5.xlsx": {
                "source_columns": ['B', 'F', 'I', 'L', 'R', 'U', 'X', 'AD', 'AJ', 'AM', 'AP', 'AS', 'AV', 'AY', 'BB'],
                "target_columns": ['B', 'D', 'E', 'F', 'H', 'I', 'J', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
            },
            "BG-ALT-S6.xlsx": {
                "source_columns": ['B', 'F', 'I', 'O', 'R', 'X', 'AA', 'AG', 'AJ', 'AM', 'AP', 'AS', 'AV'],
                "target_columns": ['B', 'D', 'E', 'G', 'H', 'J', 'K', 'M', 'N', 'O', 'P', 'Q', 'R']
            },
            
        }

        # Obtenir la configuration pour le template actuel
        config = template_configs.get(template_name)
        if not config:
            raise ValueError(f"Configuration non trouvée pour le template : {template_name}")

        source_columns = config["source_columns"]
        target_columns = config["target_columns"]
        source_start_row = 6
        target_start_row = 3

        # Copier les valeurs des cellules source vers les cellules cibles
        for source_col, target_col in zip(source_columns, target_columns):
            row_offset = 0
            for row in range(source_start_row, source_ws.max_row + 1):
                source_cell = f'{source_col}{row}'
                target_cell = f'{target_col}{target_start_row + row_offset}'

                source_value = source_ws[source_cell].value

                # Vérifier et ignorer les valeurs spécifiques
                if source_value is not None and not isinstance(source_value, str):
                    source_value = str(source_value)

                if source_value is not None and (
                    "* Attention, le total des absences" in source_value or
                    "Moyenne du groupe" in source_value
                ):
                    logging.info(f"Valeur ignorée : {source_value} (cellule {source_cell})")
                    continue

                if source_value is not None:
                    template_ws[target_cell] = source_value
                    logging.info(f"Copie de {source_value} depuis {source_cell} vers {target_cell}")
                    row_offset += 1

        # Sauvegarder le fichier mis à jour
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        updated_template_path = os.path.join(output_dir, template_name)  # Changement ici
        template_wb.save(updated_template_path)
        logging.info(f"Fichier mis à jour sauvegardé dans : {updated_template_path}")

        return updated_template_path

    except Exception as e:
        logging.error(f"Erreur lors de la copie des cellules spécifiques : {str(e)}")
        raise ValueError(f"Erreur lors de la copie des cellules spécifiques : {str(e)}")


def compare_group_code_types(template_path: str):
    """
    Compare les types des codeGroupe entre l'API et Excel
    """
    try:
        # Charger Excel
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Récupérer un codeGroupe d'Excel
        excel_code = ws["AE3"].value
        print(f"Excel codeGroupe: {excel_code} (type: {type(excel_code)})")
        
        # Récupérer un codeGroupe de l'API
        groupes = YpareoService.get_groupes()
        api_code = next(
            (groupe["codeGroupe"] 
            for groupe in groupes.values() 
            if isinstance(groupe, dict) and "codeGroupe" in groupe),
            None
        )
        print(f"API codeGroupe: {api_code} (type: {type(api_code)})")
        
        # Vérifier si les types correspondent
        print(f"Même type: {type(excel_code) == type(api_code)}")
        
    except Exception as e:
        logging.error(f"Erreur de comparaison : {str(e)}")

async def process_excel_with_template(excel_url: str, output_dir: str, prisma_template: str, user_id: str):
    """
    Processus complet : récupère le template, copie les données, et sauvegarde le fichier.
    """
    try:
        logging.info(f"Traitement du fichier Excel : {excel_url}")

        # Get template and copy cells
        template_excel_path = await get_template_from_prisma(prisma_template, output_dir)
        updated_template_path = copy_multiple_cells(excel_url, template_excel_path, output_dir)
        
        # Get Word URL from Prisma
        db = Prisma()
        await db.connect()
        config = await db.configuration.find_first(
            where={
                "excelUrl": excel_url
            }
        )
        word_url = config.wordUrl if config else None
        
        if not word_url:
            await db.disconnect()
            logging.warning("Pas d'URL Word trouvée dans la configuration")
            raise ValueError("URL du fichier Word manquante")

        # Fill template with Ypareo data and appreciations
        updated_template_path = await fill_template_with_ypareo_data(excel_url, updated_template_path, output_dir, word_url)
        
        # Lire le fichier Excel source pour obtenir le nom du groupe
        excel_response = requests.get(excel_url)
        if excel_response.status_code != 200:
            raise ValueError("Impossible de télécharger le fichier Excel source")
            
        source_excel = BytesIO(excel_response.content)
        wb = openpyxl.load_workbook(source_excel)
        ws = wb.active
        
        # Lire le nom du groupe depuis la cellule B2
        group_name = ws["B2"].value
        logging.info(f"Nom du groupe lu depuis B2: {group_name}")
        
        if not group_name:
            raise ValueError("Nom du groupe non trouvé dans la cellule B2 du fichier Excel")
            
        logging.info(f"Recherche du template pour le groupe : {group_name}")
            
        # Obtenir l'ID du template correspondant au groupe
        template_id = await get_template_id_from_group_name(db, group_name)

        # Sauvegarder l'Excel mis à jour dans Prisma
        with open(updated_template_path, 'rb') as file:
            file_content = file.read()
            # Convertir les données binaires en Base64
            file_content_base64 = base64.b64encode(file_content).decode('utf-8')
            
            # Créer l'enregistrement dans Prisma avec les données en Base64
            generated_excel = await db.generatedexcel.create({
                'data': file_content_base64,
                'userId': user_id,
                'templateId': template_id
            })
            
        await db.disconnect()
        logging.info(f"Excel sauvegardé dans Prisma avec l'ID : {generated_excel.id}")
        
        return {
            "excel_path": updated_template_path,
            "excel_id": generated_excel.id
        }

    except Exception as e:
        logging.error(f"Erreur pendant le traitement des données : {str(e)}")
        raise ValueError(f"Erreur lors du traitement du fichier Excel avec template : {str(e)}")
    
async def fill_template_with_ypareo_data(source_url: str, template_path: str, output_dir: str, word_url: str) -> str:
    """
    Remplit le template Excel avec les données Yparéo, y compris nomGroupe et etenduGroupe,
    en fonction des codeGroupe des apprenants fréquents.
    """
    try:
        
        # Télécharger et traiter le fichier Word
        word_response = requests.get(word_url)
        word_response.raise_for_status()
        temp_word_path = os.path.join(output_dir, "temp.docx")
        with open(temp_word_path, 'wb') as f:
            f.write(word_response.content)

        # Extraire les appréciations du document Word
        appreciations = {}
        doc = Document(temp_word_path)
        for table in doc.tables:
            for row in table.rows:
                if len(row.cells) >= 2:
                    nom = row.cells[0].text.strip().upper()
                    appreciation = row.cells[1].text.strip()
                    if nom and appreciation:
                        appreciations[nom] = appreciation

        # Charger les fichiers Excel
        template_wb = openpyxl.load_workbook(template_path)
        template_ws = template_wb.active
        template_name = os.path.basename(template_path)
        logging.info(f"Traitement du template : {template_name}")

        # Récupérer les données Yparéo
        frequentes = YpareoService.get_frequentes()
        groupes = YpareoService.get_groupes()
        apprenants = YpareoService.get_apprenants()
        absences = YpareoService.get_absences()
        
        # Traitement des absences
        # Traitement des absences
        absences_summary = {}
        for code_apprenant, abs_list in absences.items():
            absences_summary[str(code_apprenant)] = {  # Convertir en string pour la comparaison
                'justified': [],
                'unjustified': [],
                'delays': []
            }
            
            for absence in abs_list:
                duree = int(absence.get('duree', 0))  # Convertir en int
                if absence.get('isJustifie'):
                    absences_summary[str(code_apprenant)]['justified'].append(duree)
                elif absence.get('isRetard'):
                    absences_summary[str(code_apprenant)]['delays'].append(duree)
                else:
                    absences_summary[str(code_apprenant)]['unjustified'].append(duree)

        # Étape 1: Mapping des groupes par codeGroupe
        # Créer le mapping des groupes
        groupes_mapping = {
            str(groupe["codeGroupe"]): {
                "codeGroupe": str(groupe["codeGroupe"]),
                "nomGroupe": groupe.get("nomGroupe", ""),
                "etenduGroupe": groupe.get("etenduGroupe", "")
            }
            for groupe in groupes  # Supprimé .values() car groupes est déjà une liste
            if isinstance(groupe, dict) and "codeGroupe" in groupe
        }

        # Créer le mapping des fréquentations
        frequentation_groupe_mapping = {
            str(frequentation.get("codeApprenant", "")): str(frequentation.get("codeGroupe", ""))
            for frequentation in frequentes
            if isinstance(frequentation, dict)
        }
        
        # Créer le mapping complet des apprenants avec leurs données de groupe
        apprenant_mapping = {
            f"{a['nomApprenant'].strip().upper()} {a['prenomApprenant'].strip().upper()}": {
                "codeApprenant": str(a.get("codeApprenant", "")),
                "dateNaissance": str(a.get("dateNaissance", "")),
                "site": str(a.get("inscriptions", [{}])[0].get("site", {}).get("nomSite", "")),
                **groupes_mapping.get(
                    frequentation_groupe_mapping.get(
                        str(a.get("codeApprenant", "")), 
                        ""
                    ), 
                    {"codeGroupe": "", "nomGroupe": "", "etenduGroupe": ""}
                )
            }
            for a in apprenants if isinstance(a, dict)
        }
        
        # Configuration des colonnes pour chaque template
        template_configs = {
            "BG-TP-S1.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "AC",
                "site": "AD",
                "code_groupe": "AE",
                "nom_groupe": "AF",
                "etendu_groupe": "AG",
                "abs_justified": "AH",
                "abs_unjustified": "AI",
                "abs_delays": "AJ",
                "appreciation": "AK"
            },
            "BG-TP-S2.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "F",
                "site": "G",
                "code_groupe": "H",
                "nom_groupe": "I",
                "etendu_groupe": "J",
                "abs_justified": "K",
                "abs_unjustified": "L",
                "abs_delays": "M",
                "appreciation": "N"
            },
            "BG-TP-S3.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "V",
                "site": "W",
                "code_groupe": "X",
                "nom_groupe": "Y",
                "etendu_groupe": "Z",
                "abs_justified": "AA",
                "abs_unjustified": "AB",
                "abs_delays": "AC",
                "appreciation": "AD"
            },
            "BG-TP-S4.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "F",
                "site": "G",
                "code_groupe": "H",
                "nom_groupe": "I",
                "etendu_groupe": "J",
                "abs_justified": "K",
                "abs_unjustified": "L",
                "abs_delays": "M",
                "appreciation": "N"
            },
            "BG-TP-S5.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "Z",
                "site": "AA",
                "code_groupe": "AB",
                "nom_groupe": "AC",
                "etendu_groupe": "AD",
                "abs_justified": "AE",
                "abs_unjustified": "AF",
                "abs_delays": "AG",
                "appreciation": "AH"
            },
            "BG-TP-S6.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "G",
                "site": "H",
                "code_groupe": "I",
                "nom_groupe": "J",
                "etendu_groupe": "K",
                "abs_justified": "L",
                "abs_unjustified": "M",
                "abs_delays": "N",
                "appreciation": "O"
            },
            "BG-ALT-S1.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "U",
                "site": "V",
                "code_groupe": "W",
                "nom_groupe": "X",
                "etendu_groupe": "Y",
                "abs_justified": "Z",
                "abs_unjustified": "AA",
                "abs_delays": "AB",
                "appreciation": "AC"
            },
            "BG-ALT-S2.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "V",
                "site": "W",
                "code_groupe": "X",
                "nom_groupe": "Y",
                "etendu_groupe": "Z",
                "abs_justified": "AA",
                "abs_unjustified": "AB",
                "abs_delays": "AC",
                "appreciation": "AD"
            },
            "BG-ALT-S3.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "T",
                "site": "U",
                "code_groupe": "V",
                "nom_groupe": "W",
                "etendu_groupe": "X",
                "abs_justified": "Y",
                "abs_unjustified": "Z",
                "abs_delays": "AA",
                "appreciation": "AB"
            },
            "BG-ALT-S4.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "T",
                "site": "U",
                "code_groupe": "V",
                "nom_groupe": "W",
                "etendu_groupe": "X",
                "abs_justified": "Y",
                "abs_unjustified": "Z",
                "abs_delays": "AA",
                "appreciation": "AB"
            },
            "BG-ALT-S5.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "U",
                "site": "v",
                "code_groupe": "W",
                "nom_groupe": "X",
                "etendu_groupe": "Y",
                "abs_justified": "Z",
                "abs_unjustified": "AA",
                "abs_delays": "AB",
                "appreciation": "AC"
            },
            "BG-ALT-S6.xlsx": {
                "code_apprenant": "A",
                "date_naissance": "S",
                "site": "T",
                "code_groupe": "U",
                "nom_groupe": "V",
                "etendu_groupe": "W",
                "abs_justified": "X",
                "abs_unjustified": "Y",
                "abs_delays": "Z",
                "appreciation": "AA"
            },
        }

        config = template_configs.get(template_name)
        if not config:
            logging.error(f"Template name: {template_name}")
            logging.error(f"Available configs: {list(template_configs.keys())}")
            raise ValueError(f"Configuration non trouvée pour le template : {template_name}")

        # Vérifier que toutes les clés nécessaires sont présentes
        required_keys = ['date_naissance', 'site', 'code_groupe', 'nom_groupe', 'etendu_groupe', 
                        'abs_justified', 'abs_unjustified', 'abs_delays', 'appreciation']
        missing_keys = [key for key in required_keys if key not in config]
        if missing_keys:
            raise ValueError(f"Configuration incomplète pour {template_name}. Clés manquantes : {missing_keys}")

        
        #
        # Ajouter ces logs de débogage après la récupération des données Yparéo
        logging.info(f"Template utilisé : {template_name}")
        logging.info(f"Configuration utilisée : {config}")
        logging.info(f"Nombre d'apprenants trouvés : {len(apprenants)}")
        logging.info(f"Nombre de groupes trouvés : {len(groupes)}")
        logging.info(f"Nombre de fréquentations trouvées : {len(frequentes)}")

        # Dans la boucle de remplissage, ajouter des logs détaillés
        for row in range(3, template_ws.max_row + 1):
            template_nom_prenom = template_ws[f"B{row}"].value
            if template_nom_prenom:
                normalized_nom_prenom = template_nom_prenom.strip().upper()
                logging.info(f"Traitement de l'apprenant : {normalized_nom_prenom}")
                
                if normalized_nom_prenom in apprenant_mapping:
                    apprenant_data = apprenant_mapping[normalized_nom_prenom]
                    code_apprenant = apprenant_data["codeApprenant"]
                    logging.info(f"Données trouvées pour {normalized_nom_prenom}: {apprenant_data}")

                    # Remplir les données selon la configuration du template
                    try:
                        if "code_apprenant" in config:
                            template_ws[f"{config['code_apprenant']}{row}"].value = code_apprenant
                            template_ws[f"{config['date_naissance']}{row}"].value = apprenant_data["dateNaissance"]
                            template_ws[f"{config['site']}{row}"].value = apprenant_data["site"]
                            template_ws[f"{config['code_groupe']}{row}"].value = apprenant_data["codeGroupe"]
                            template_ws[f"{config['nom_groupe']}{row}"].value = apprenant_data["nomGroupe"]
                            template_ws[f"{config['etendu_groupe']}{row}"].value = apprenant_data["etenduGroupe"]
                            logging.info(f"Données remplies avec succès pour {normalized_nom_prenom} à la ligne {row}")
                        
                        # Remplir les absences
                        if code_apprenant in absences_summary:
                            abs_info = absences_summary[code_apprenant]
                            template_ws[f"{config['abs_justified']}{row}"].value = convert_minutes_to_hours_and_minutes(sum(abs_info['justified']))
                            template_ws[f"{config['abs_unjustified']}{row}"].value = convert_minutes_to_hours_and_minutes(sum(abs_info['unjustified']))
                            template_ws[f"{config['abs_delays']}{row}"].value = convert_minutes_to_hours_and_minutes(sum(abs_info['delays']))

                        # Remplir l'appréciation
                        if normalized_nom_prenom in appreciations:
                            template_ws[f"{config['appreciation']}{row}"].value = appreciations[normalized_nom_prenom]
                            logging.info(f"Appréciation ajoutée pour {normalized_nom_prenom}")
                    except Exception as e:
                        logging.error(f"Erreur lors du remplissage des données pour {normalized_nom_prenom}: {str(e)}")

        # Sauvegarder le fichier
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        updated_template_path = os.path.join(output_dir, "updated_excel.xlsx")
        template_wb.save(updated_template_path)
        logging.info(f"Fichier template mis à jour sauvegardé à : {updated_template_path}")
        return updated_template_path

    except Exception as e:
        logging.error(f"Erreur lors du remplissage des données dans le template : {str(e)}")
        raise ValueError(f"Erreur lors du remplissage des données dans le template : {str(e)}")

    
async def match_template_and_get_word(updated_excel_path):
    """
    Compare le fichier Excel mis à jour avec les templates et retourne le modèle Word approprié.
    """
    try:
        logging.info("Début de la comparaison des templates...")
        
        # Charger le fichier Excel mis à jour
        updated_wb = openpyxl.load_workbook(updated_excel_path)
        updated_ws = updated_wb.active

        # Récupérer les valeurs des cellules à comparer pour BG-ALT-S3
        cells_to_compare_s3 = ['C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1']
        updated_values_s3 = [str(updated_ws[cell].value or '').strip() for cell in cells_to_compare_s3]

        # Récupérer les valeurs des cellules à comparer pour BG-ALT-S2
        cells_to_compare_s2 = ['C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 'U1']
        updated_values_s2 = [str(updated_ws[cell].value or '').strip() for cell in cells_to_compare_s2]

        # Récupérer les templates depuis Prisma
        template_s3_data = await fetch_template_from_prisma("BG-ALT-S3.xlsx")
        template_s2_data = await fetch_template_from_prisma("BG-ALT-S2.xlsx")

        template_s3_wb = openpyxl.load_workbook(BytesIO(template_s3_data))
        template_s2_wb = openpyxl.load_workbook(BytesIO(template_s2_data))

        template_s3_ws = template_s3_wb.active
        template_s2_ws = template_s2_wb.active

        # Comparer les valeurs avec BG-ALT-S3
        template_s3_values = [str(template_s3_ws[cell].value or '').strip() for cell in cells_to_compare_s3]
        matches_s3 = updated_values_s3 == template_s3_values

        # Comparer les valeurs avec BG-ALT-S2
        template_s2_values = [str(template_s2_ws[cell].value or '').strip() for cell in cells_to_compare_s2]
        matches_s2 = updated_values_s2 == template_s2_values

        # Créer le dossier temp s'il n'existe pas
        temp_dir = "./temp"
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir)

        # Déterminer quel template utiliser
        if matches_s3:
            logging.info("Template correspondant trouvé: BG-ALT-S3.xlsx")
            word_data = await fetch_template_from_prisma("modeleBGALT3.docx")
            word_path = os.path.join(temp_dir, "modeleBGALT3.docx")
            template_name = "modeleBGALT3.docx"
            ects_template = "BG_ALT_3"
        elif matches_s2:
            logging.info("Template correspondant trouvé: BG-ALT-S2.xlsx")
            word_data = await fetch_template_from_prisma("modeleBGALT2.docx")
            word_path = os.path.join(temp_dir, "modeleBGALT2.docx")
            template_name = "modeleBGALT2.docx"
            ects_template = "BG_ALT_2"  # Make sure this matches exactly with the ECTS data key
        else:
            logging.warning("Aucune correspondance trouvée avec les templates")
            raise ValueError("Impossible de déterminer le template à utiliser")

        # Sauvegarder le modèle Word
        with open(word_path, 'wb') as f:
            f.write(word_data)
            
        logging.info(f"Modèle Word {template_name} sauvegardé dans {word_path}")
        return {
            "word_path": word_path,
            "template_name": template_name,
            "ects_template": ects_template
        }

    except Exception as e:
        logging.error(f"Erreur lors de la comparaison des templates: {str(e)}")
        raise
